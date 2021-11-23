import tkinter
import openpyxl


from functools import partial

#Text用不到
def TextPrint():
    """text.delete(1.0,"end")
    text.insert("end",str1)
    text.insert("end","\n")
    text.insert("end",str2)
    text.insert("end", "\n")
    text.insert("end", str3)
    text.insert("end", "\n")
    text.insert("end", str4)
    text.insert("end", "\n")
    text.insert("end", str5)"""
    return

def Click11():
    if str1[0]=="0":
        button11['bg']='orange'
        str1[0]="1"
    else:
        button11['bg'] = 'gray'
        str1[0]="0"
    TextPrint()
def Click12():
    if str1[1] == "0":
        button12['bg'] = 'orange'
        str1[1] = "1"
    else:
        button12['bg'] = 'gray'
        str1[1] = "0"
    TextPrint()
def Click13():
    if str1[2] == "0":
        button13['bg'] = 'orange'
        str1[2] = "1"
    else:
        button13['bg'] = 'gray'
        str1[2] = "0"
    TextPrint()
def Click14():
    if str1[3] == "0":
        button14['bg'] = 'orange'
        str1[3] = "1"
    else:
        button14['bg'] = 'gray'
        str1[3] = "1"
    TextPrint()
def Click15():
    if str1[4] == "0":
        button15['bg'] = 'orange'
        str1[4] = "1"
    else:
        button15['bg'] = 'gray'
        str1[4] = "0"
    TextPrint()

def Click21():
    if str2[0]=="0":
        button21['bg'] = 'orange'
        str2[0]="1"
    else:
        button21['bg'] = 'gray'
        str2[0]="0"
    TextPrint()
def Click22():
    if str2[1] == "0":
        button22['bg'] = 'orange'
        str2[1] = "1"
    else:
        button22['bg'] = 'gray'
        str2[1] = "0"
    TextPrint()
def Click23():
    if str2[2] == "0":
        button23['bg'] = 'orange'
        str2[2] = "1"
    else:
        button23['bg'] = 'gray'
        str2[2] = "0"
    TextPrint()
def Click24():
    if str2[3] == "0":
        button24['bg'] = 'orange'
        str2[3] = "1"
    else:
        button24['bg'] = 'gray'
        str2[3] = "0"
    TextPrint()
def Click25():
    if str2[4] == "0":
        button25['bg'] = 'orange'
        str2[4] = "1"
    else:
        button25['bg'] = 'gray'
        str2[4] = "0"
    TextPrint()

def Click31():
    if str3[0]=="0":
        button31['bg'] = 'orange'
        str3[0]="1"
    else:
        button31['bg'] = 'gray'
        str3[0]="0"
    TextPrint()
def Click32():
    if str3[1] == "0":
        button32['bg'] = 'orange'
        str3[1] = "1"
    else:
        button32['bg'] = 'gray'
        str3[1] = "0"
    TextPrint()
def Click33():
    if str3[2] == "0":
        button33['bg'] = 'orange'
        str3[2] = "1"
    else:
        button33['bg'] = 'gray'
        str3[2] = "0"
    TextPrint()
def Click34():
    if str3[3] == "0":
        button34['bg'] = 'orange'
        str3[3] = "1"
    else:
        button34['bg'] = 'gray'
        str3[3] = "0"
    TextPrint()
def Click35():
    if str3[4] == "0":
        button35['bg'] = 'orange'
        str3[4] = "1"
    else:
        button35['bg'] = 'gray'
        str3[4] = "0"
    TextPrint()

def Click41():
    if str4[0]=="0":
        button41['bg'] = 'orange'
        str4[0]="1"
    else:
        button41['bg'] = 'gray'
        str4[0]="0"
    TextPrint()
def Click42():
    if str4[1] == "0":
        button42['bg'] = 'orange'
        str4[1] = "1"
    else:
        button42['bg'] = 'gray'
        str4[1] = "0"
    TextPrint()
def Click43():
    if str4[2] == "0":
        button43['bg'] = 'orange'
        str4[2] = "1"
    else:
        button43['bg'] = 'gray'
        str4[2] = "0"
    TextPrint()
def Click44():
    if str4[3] == "0":
        button44['bg'] = 'orange'
        str4[3] = "1"
    else:
        button44['bg'] = 'gray'
        str4[3] = "0"
    TextPrint()
def Click45():
    if str4[4] == "0":
        button45['bg'] = 'orange'
        str4[4] = "1"
    else:
        button45['bg'] = 'gray'
        str4[4] = "0"
    TextPrint()

def Click51():
    if str5[0]=="0":
        button51['bg'] = 'orange'
        str5[0]="1"
    else:
        button51['bg'] = 'gray'
        str5[0]="0"
    TextPrint()
def Click52():
    if str5[1] == "0":
        button52['bg'] = 'orange'
        str5[1] = "1"
    else:
        button52['bg'] = 'gray'
        str5[1] = "0"
    TextPrint()
def Click53():
    if str5[2] == "0":
        button53['bg'] = 'orange'
        str5[2] = "1"
    else:
        button53['bg'] = 'gray'
        str5[2] = "0"
    TextPrint()
def Click54():
    if str5[3] == "0":
        button54['bg'] = 'orange'
        str5[3] = "1"
    else:
        button54['bg'] = 'gray'
        str5[3] = "0"
    TextPrint()
def Click55():
    if str5[4] == "0":
        button55['bg'] = 'orange'
        str5[4] = "1"
    else:
        button55['bg'] = 'gray'
        str5[4] = "0"
    TextPrint()

#清空输入
def Refresh():
    for i in range(0,5):
        str1[i] = '0'
        str2[i] = '0'
        str3[i] = '0'
        str4[i] = '0'
        str5[i] = '0'
    button11['bg'] = 'gray'
    button12['bg'] = 'gray'
    button13['bg'] = 'gray'
    button14['bg'] = 'gray'
    button15['bg'] = 'gray'

    button21['bg'] = 'gray'
    button22['bg'] = 'gray'
    button23['bg'] = 'gray'
    button24['bg'] = 'gray'
    button25['bg'] = 'gray'

    button31['bg'] = 'gray'
    button32['bg'] = 'gray'
    button33['bg'] = 'gray'
    button34['bg'] = 'gray'
    button35['bg'] = 'gray'

    button41['bg'] = 'gray'
    button42['bg'] = 'gray'
    button43['bg'] = 'gray'
    button44['bg'] = 'gray'
    button45['bg'] = 'gray'

    button51['bg'] = 'gray'
    button52['bg'] = 'gray'
    button53['bg'] = 'gray'
    button54['bg'] = 'gray'
    button55['bg'] = 'gray'

#录入信息
def Save():
    data_output = int(entry.get())
    sheet.append([data_output,
                  int(str1[0]), int(str1[1]), int(str1[2]), int(str1[3]), int(str1[4]),
                  int(str2[0]), int(str2[1]), int(str2[2]), int(str2[3]), int(str2[4]),
                  int(str3[0]), int(str3[1]), int(str3[2]), int(str3[3]), int(str3[4]),
                  int(str4[0]), int(str4[1]), int(str4[2]), int(str4[3]), int(str4[4]),
                  int(str5[0]), int(str5[1]), int(str5[2]), int(str5[3]), int(str5[4])
                  ])
    entry.delete(0,"end")
    wb.save('ceshiji.xlsx')
    return

#朴素贝叶斯预测
def Test():
    PCount = [0.0]*10
    PC = [0.0]*10
    PResult = [1.0] * 10
    PTest = [0.0]*2
    PX = [[[0.0 for i in range(2)] for j in range(25)] for k in range(10)]
    #PX[十个输出结果][二十五个属性][零一两种结果]
    #其中C表示输出结果0，1,2......9
    for i in range(0,10):
        for j in range(1,sheetlen+1):
            if sheet.cell(j,1).value == i:
                PCount[i]+=1

    for i in range(0,10):
        PC[i]=PCount[i]/sheetlen
        #print(PC[i])


    for i in range(0,10):#10个输出结果

        for j in range(0,25):#25个属性

            for k in range(1,sheetlen+1):#统计零和一各出现多少
                #统计第k个属性中，值为0/1且输出值为i的数量
                if sheet.cell(k,j+2).value==0 and sheet.cell(k,1).value==i:
                    PX[i][j][0]+=1
                elif sheet.cell(k,j+2).value==1 and sheet.cell(k,1).value==i :
                    PX[i][j][1]+=1

            PX[i][j][0] = PX[i][j][0]/PCount[i]
            PX[i][j][1] = PX[i][j][1]/PCount[i]
            #print("p0",i," ",j,"等于",PX[i][j][0])
            #print("p1", i, " ", j, "等于", PX[i][j][1])
            #print("满足输出值",i,"且","第",j,"个属性中出现零的个数是",PX[i][j][0])

    tlist = [0 for i in range(0,25)]

    j=0
    for i in range(0,4+1):
        tlist[i]=int(str1[j])
        j+=1
    j = 0
    for i in range(5,9+1):
        tlist[i]=int(str2[j])
        j+=1
    j=0
    for i in range(10,14+1):
        tlist[i]=int(str3[j])
        j+=1
    j=0
    for i in range(15,19+1):
        tlist[i]=int(str4[j])
        j+=1
    j=0
    for i in range(20,24+1):
        tlist[i]=int(str5[j])
        j+=1

    #print(tlist)


    PResult[0] = PResult[0] * PC[0]
    #print("PResult=",PResult[0])

    #print(PResult[0])


    for i in range(0,10):
        PResult[i] = PResult[i]*PC[i]
        for j in range(0,25):
            if (tlist[j]==0):
                PResult[i] = PResult[i]*PX[i][j][0]
            else:
                PResult[i] = PResult[i]*PX[i][j][1]
        #print(PResult[i])


    Max = 0
    Maxi = 0
    for i in range(0,10):
        print(PResult[i])
        if (PResult[i]>Max):
            Max = PResult[i];
            Maxi = i;

    entry.delete(0,"end")
    entry.insert("end",Maxi)




#读取Excel
wb = openpyxl.load_workbook('ceshiji.xlsx')
sheet = wb['Sheet1']
sheetlen = sheet.max_row
sheetwidth = sheet.max_column
#界面
root = tkinter.Tk() # 初始化Tk()
root.title("framework")
root.resizable(0, 0)

button_bg = 'orange'
math_sign_bg = 'DarkTurquoise'
cal_output_bg = 'YellowGreen'
button_active_bg = 'gray'


str1 = "00000"
str1 = list(str1)
str2 = "00000"
str2 = list(str2)
str3 = "00000"
str3 = list(str3)
str4 = "00000"
str4 = list(str4)
str5 = "00000"
str5 = list(str5)


#text = tkinter.Text(root,font=1)
#text.grid(row=0, column=0, columnspan=5, padx=10, pady=10)
entry = tkinter.Entry(root)
entry.grid(row=0, column=0, columnspan=5,padx=10, pady=10)
TextPrint()


button11 = tkinter.Button(root,text="11",bg='gray',command=Click11)
button11.grid(row=1,column=0,ipadx=10,ipady=10,padx=10)
button12 = tkinter.Button(root, text="12",bg='gray', command=Click12)
button12.grid(row=1, column=1,ipadx=10, ipady=10,padx=10)
button13 = tkinter.Button(root,text="13",bg='gray',command=Click13)
button13.grid(row=1,column=2,ipadx=10,ipady=10,padx=10)
button14 = tkinter.Button(root,text="14",bg='gray',command=Click14)
button14.grid(row=1,column=3,ipadx=10,ipady=10,padx=10)
button15 = tkinter.Button(root,text="15",bg='gray',command=Click15)
button15.grid(row=1,column=4,ipadx=10,ipady=10,padx=10)

button21 = tkinter.Button(root,text="21",bg='gray',command=Click21)
button21.grid(row=2,column=0,ipadx=10,ipady=10,padx=10)
button22 = tkinter.Button(root,text="22",bg='gray', command=Click22)
button22.grid(row=2, column=1,ipadx=10, ipady=10,padx=10)
button23 = tkinter.Button(root,text="23",bg='gray',command=Click23)
button23.grid(row=2,column=2,ipadx=10,ipady=10,padx=10)
button24 = tkinter.Button(root,text="24",bg='gray',command=Click24)
button24.grid(row=2,column=3,ipadx=10,ipady=10,padx=10)
button25 = tkinter.Button(root,text="25",bg='gray',command=Click25)
button25.grid(row=2,column=4,ipadx=10,ipady=10,padx=10)

button31 = tkinter.Button(root,text="31",bg='gray',command=Click31)
button31.grid(row=3,column=0,ipadx=10,ipady=10,padx=10)
button32 = tkinter.Button(root,text="32",bg='gray', command=Click32)
button32.grid(row=3, column=1,ipadx=10, ipady=10,padx=10)
button33 = tkinter.Button(root,text="33",bg='gray',command=Click33)
button33.grid(row=3,column=2,ipadx=10,ipady=10,padx=10)
button34 = tkinter.Button(root,text="34",bg='gray',command=Click34)
button34.grid(row=3,column=3,ipadx=10,ipady=10,padx=10)
button35 = tkinter.Button(root,text="35",bg='gray',command=Click35)
button35.grid(row=3,column=4,ipadx=10,ipady=10,padx=10)

button41 = tkinter.Button(root,text="41",bg='gray',command=Click41)
button41.grid(row=4,column=0,ipadx=10,ipady=10,padx=10)
button42 = tkinter.Button(root,text="42",bg='gray', command=Click42)
button42.grid(row=4, column=1,ipadx=10, ipady=10,padx=10)
button43 = tkinter.Button(root,text="43",bg='gray',command=Click43)
button43.grid(row=4,column=2,ipadx=10,ipady=10,padx=10)
button44 = tkinter.Button(root,text="44",bg='gray',command=Click44)
button44.grid(row=4,column=3,ipadx=10,ipady=10,padx=10)
button45 = tkinter.Button(root,text="45",bg='gray',command=Click45)
button45.grid(row=4,column=4,ipadx=10,ipady=10,padx=10)

button51 = tkinter.Button(root,text="51",bg='gray',command=Click51)
button51.grid(row=5,column=0,ipadx=10,ipady=10,padx=10)
button52 = tkinter.Button(root,text="52",bg='gray', command=Click52)
button52.grid(row=5, column=1,ipadx=10, ipady=10,padx=10)
button53 = tkinter.Button(root,text="53",bg='gray',command=Click53)
button53.grid(row=5,column=2,ipadx=10,ipady=10,padx=10)
button54 = tkinter.Button(root,text="54",bg='gray',command=Click54)
button54.grid(row=5,column=3,ipadx=10,ipady=10,padx=10)
button55 = tkinter.Button(root,text="55",bg='gray',command=Click55)
button55.grid(row=5,column=4,ipadx=10,ipady=10,padx=10)

button6 = tkinter.Button(root,text="存入",command=Save)
button6.grid(row=6,column=0,padx=10,pady=10)
button7 = tkinter.Button(root,text="测试",command=Test)
button7.grid(row=6,column=1,padx=10,pady=10)
button8 = tkinter.Button(root,text="恢复",command=Refresh)
button8.grid(row=6,column=2,padx=10,pady=10)

root.mainloop()
