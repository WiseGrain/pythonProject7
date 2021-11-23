import openpyxl

wb = openpyxl.load_workbook('test.xlsx')
sheet = wb['aa']


PC = [0]*3

PX = [[[0.0 for i in range(2)] for j in range(10)] for k in range(25)]

PX[0][0][1]=1
print(PX)

"""l = sheet.max_row
print(l)"""

"""
c = sheet['B11'].value
print(c)
print(sheet.cell(row=20,column=2).value)

for i in range(1,20):
    print(sheet.cell(row=i,column=2).value)
    
"""#读取操作

"""
sheet.append([1,2,3,4,5,6])


wb.save('test.xlsx')
"""
